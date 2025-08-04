import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# Path to your Excel file (where it logs checkouts)
excel_path = r"C:\Users\amandac\Western Building Group\FileShare - Documents\Lisa & Amanda\Amanda - AI\AZ Part List.xlsx"
sheet_name = "Checkout"

# --- UI: Title ---
st.title("Inventory Checkout Log - Arizona")

# --- UI: Input Fields ---
name = st.text_input("Your Name")
item = st.selectbox("Select Item", [
    "DB-1/8", "DB-#30", "DB-3/16", "DB-1/4", "DB-3/16 SDS", "DB-1/4 SDS", "CSB-7 1/4", "CSB-6 1/2", "CSB-5 3/8",
    "CSB-7 1/4 AL", "CSB-6 1/2 AL", "CW-4 1/2", "CW-5", "RZBLD", "BSB-44 7/8", "RSB-6", "RSB-12", "JSB-24T", "JSB-TS",
    "ACM-RB", "BTH-3", "BTH-6", "PBT #2R", "PBT #3", "SBT #2", "SBT #3", "TBT-T25", "DCB", "ND-1/4 x6", "ND-5/16 x6",
    "ND-3/8 x6", "STLN", "RK-Jw/s", "RK-PT", "STPL-T-50", "#8-1 ST", "#8-3/4 ST", "#8-1 AP", "#8-3/4 AP", "#10-1 PN-T2",
    "1 HWH-WOOD", "2 HWH-WOOD", "1 PN-T3", "1 1/4 PN T-17", "1 1/4 PN T-3", "2PN-T3", "3 PN-T3", "1 HWH-T3",
    "1 1/2 HWH-T3", "2 HWH-T3", "2 1/2 HWH-T3", "3 HWH-T3", "1 1/4 T5", "1 1/2 T5", "3 T5", "1 1/4 T5 TORX",
    "RVT #44 AL", "RVT #46 AL", "RVT #44 SS", "RVT #46 SS", "3/16 NAILIN", "1 1/4 TAP HWH 1/4", "1 3/4 TAP HWH 1/4",
    "2 1/4 TAP HWH 1/4", "BKR-7/8 OPEN", "SG 795 WHT", "SG 795 CHR", "SG 795 GRY", "SG 795 BLK", "SG 795 SND",
    "SG 795 CHM", "SGT-YELLOW", "SGT-GREEN", "SGT-RNG", "SB 1/16", "SR 1/8", "SB 1/4"
])
quantity = st.number_input("Quantity", min_value=1, step=1)
unit_type = st.selectbox("Quantity Type", ["Individual piece(s)", "Bag", "Box"])

job_options = [
    "21048 - TSMC Fab 21 - Phoenix, AZ", "21064 - Carvana Corporate Campus", "21066 - Maya Hotel",
    "22004 - QTS PHX2 DC1", "22006 - Historic Post Office", "22018 - Gilbert NWTP South Reservoir Improvements",
    "22025 - Caesars Republic Hotel", "22044 - Vantage Data Center AZ 12 & 13",
    "22049 - Tempe Municipal Operations Center - GMP 3", "22052 - The McKinley",
    "22060 - Prologis Loop 303 Goodyear Building 1", "23001 - 601 N Central",
    "23002 - Mesa Police Evidence Storage", "23006 - NTT PH2-PH3", "23017 - COE Flight Engineering Lab Complex Edwards AFB",
    "23019 - Elliot Mesa Commerce Center", "23020 - QTS PHX2 DC5", "23021 - Hopi HCC Outpatient and Emergency Expansion",
    "23023 - Fry's Store 655 - Gilbert", "23025 - Goodyear Civic Square Building 3B & 4",
    "23028 - City of Mesa Northeast Public Safety Facility", "23038 - Expansion & Modernization Project at the San Luis I Land Port of Entry",
    "23039 - Esplanade Renovation - Tower 01 and 02", "23042 - Thunderbird Reservoir",
    "23046 - ASU - West Campus Academic Building", "23053 - Queen Creek Rec. & Aquatic Center",
    "23054 - Andretti's Indoor Karting Glendale, AZ", "23055 - QTS PHX2 DC2 36MW", "23056 - Phoenix Fire Station #62",
    "23059 - Ellsworth Ranch Central Amenity", "23060 - QTS PHX2 DC-3", "23063 - Electric Pickle Tempe II",
    "23064 - ASU Polytechnic Research and Educational Building", "23065 - 301 Maricopa County Restack - Phoenix, AZ",
    "23067 - South Pier Towers 1, 2, and 3", "23069 - Marana Community and Aquatic Facility",
    "24001 - Prologis Loop 303 Building 2", "24002 - Abrazo Buckeye MOB", "24004 - Yuma Administration Services Building (aka: YUCO Admin)",
    "24006 - NRS Logistics", "24007 - Surprise Fire Station #309", "24008 - Xnrgy", "24009 - Surprise Oasis Aquatic Complex",
    "24010 - Buckeye Commons Buildings A, B, C, D", "24012 - Tempe Municipal Operations Center Ph 2",
    "24013 - Papago Golf House", "24014 - Intel OC43", "24015 - Casa Grande Fire Station #503",
    "24017 - Lucid BIW Slab Replacements", "24018 - Floreo at Teravalis", "24020 - Avalon Crossing Pavilion",
    "24021 - QTS 23 PHX2 - DC1 PHXMercury EFO", "24025 - 1020 Apache", "24028 - Arizona Dignity Health R&R",
    "24031 - Sun Life Family Health Medical Office Building", "24032 - QTS PHX3 DC-14", "24034 - Hayden's Ferry Restaurant",
    "24036 - Mountain Park Health Center", "24040 - Abrazo Medical Office", "24042 - EdgeConnex PHX11",
    "24043 - Buckeye 911 Call Center", "24046 - QTS PHX3 DC-13", "24047 - Win Aviation Hangar",
    "24048 - Evie's Pavilion Expansion", "24049 - YPG Ready Building", "24050 - Casino Del Sol",
    "24051 - Uptown Mall Parking Garage", "24052 - Willscot HQ", "24058 - Fender Office Building",
    "25002 - TSMC FAB21 DB", "25006 - Nuvision Credit Union", "25007 - Sprouts HQ Design Assist",
    "25008 - NAH Orthopedic Surgery Center", "25009 - NTT PH4", "25010 - Intel - Cleanlink",
    "25011 - Astria", "25012 - Verrado Marketplace", "25013 - Hayden's Ferry Façade Renovation CO#1",
    "25016 - PSHIA Terminal 3 North Concourse 2", "25017 - PSHIA Terminal 3 South Gates",
    "25018 - Remi Hotel", "25022 - Tempe Community Action Agency", "25024 - Behrhorst Residence Re-Roof",
    "25027 - City North Dual Brand Hotel", "25028 - QTS PHX3 DC-2", "25029 - QTS PHX3 DC-9",
    "25030 - PHX 065", "25033 - Heroes Regional Park Library Expansion", "25034 - NTT PH5",
    "25035 - Desert Mountain MSA #2", "25037 - TSMC Phase 3 GMP CSA"
]

job_selected = st.selectbox("Select Job", job_options)

checkout_date = st.date_input("Checkout Date", value=datetime.today())
checkout_time = st.time_input("Checkout Time", value=datetime.now().time())

# --- Submission Logic ---
if st.button("Submit"):
    if name.strip() == "":
        st.warning("Please enter your name.")
    else:
        new_row = {
            "Timestamp": f"{checkout_date.strftime('%Y-%m-%d')} {checkout_time.strftime('%I:%M %p')}",
            "Name": name,
            "Job": job_selected,
            "Item": item,
            "Quantity": quantity,
            "Quantity Type": unit_type
        }

        # Write to Excel
        book = load_workbook(excel_path)
        sheet = book[sheet_name]
        next_row = sheet.max_row + 1

        writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        df = pd.DataFrame([new_row])
        write_header = next_row == 1
        df.to_excel(writer, sheet_name=sheet_name, startrow=next_row, index=False, header=write_header)
        writer.close()

        st.success(f"{quantity} x {item} checked out by {name} on {checkout_date} at {checkout_time}")
        st.write("✅ Data written:", df)
